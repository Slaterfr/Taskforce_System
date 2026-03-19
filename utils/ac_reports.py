"""
AC Report Generation and Title Rewards Calculator
Generates Excel reports and calculates title rewards automatically
"""

from datetime import datetime
from typing import Dict, List, Optional
import io

class ACReportGenerator:
    """Generates AC reports and calculates title rewards"""
    
    def __init__(self, ac_period, members_progress):
        self.ac_period = ac_period
        self.members_progress = members_progress
        self.title_winners = {}
        
    def generate_excel_report(self) -> io.BytesIO:
        """Generate Excel report for AC period"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel reports. Install with: pip install openpyxl")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AC Report"
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Activity Check Report - {self.ac_period.period_name}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Period info
        ws.merge_cells('A2:G2')
        period_cell = ws['A2']
        period_cell.value = f"{self.ac_period.start_date.strftime('%B %d, %Y')} - {self.ac_period.end_date.strftime('%B %d, %Y')}"
        period_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Discord Username', 'Rank', 'Quota', 'Points Earned', 'Status', 'Activities', 'Notes']
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row_num = 5
        for progress in sorted(self.members_progress, key=lambda x: x['member'].discord_username):
            member = progress['member']
            
            # Determine status
            if progress['is_protected']:
                status = "PROTECTED (IA)"
                status_color = "ADD8E6"  # Light blue
            elif progress['points'] >= progress['quota']:
                status = "PASSED"
                status_color = "90EE90"  # Light green
            else:
                status = "FAILED"
                status_color = "FFB6C1"  # Light red
            
            # Get activity breakdown
            activities_text = self._get_activity_breakdown(progress['recent_activities'])
            
            # Add row
            ws.append([
                member.discord_username,
                member.current_rank,
                progress['quota'],
                round(progress['points'], 1),
                status,
                len(progress['recent_activities']),
                progress.get('notes', '')
            ])
            
            # Style status cell
            status_cell = ws.cell(row=row_num, column=5)
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            status_cell.alignment = Alignment(horizontal='center')
            
            row_num += 1
        
        # Summary statistics
        ws.append([])
        ws.append(['SUMMARY'])
        ws.cell(row=row_num + 1, column=1).font = Font(bold=True)
        
        total = len(self.members_progress)
        passed = sum(1 for p in self.members_progress if not p['is_protected'] and p['points'] >= p['quota'])
        failed = sum(1 for p in self.members_progress if not p['is_protected'] and p['points'] < p['quota'])
        protected = sum(1 for p in self.members_progress if p['is_protected'])
        
        ws.append(['Total Members:', total])
        ws.append(['Passed:', passed])
        ws.append(['Failed:', failed])
        ws.append(['Protected (IA):', protected])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _get_activity_breakdown(self, activities) -> str:
        """Get activity breakdown as text"""
        activity_counts = {}
        for activity in activities:
            activity_type = activity.activity_type
            activity_counts[activity_type] = activity_counts.get(activity_type, 0) + 1
        
        return ", ".join([f"{k}: {v}" for k, v in activity_counts.items()])
    
    def calculate_title_rewards(self, all_activities) -> Dict[str, Dict]:
        """
        Calculate title reward winners based on activity counts
        
        Title Award Schedule:
        - Host with the Most (HWTM): Awarded EVERY period based on that period's stats only
        - Other Titles: Awarded at END OF MONTH only, using accumulated stats from all periods in month
        
        Title Requirements:
        - Host with the Most: Most events (Training + Raid + Patrol) in single period (min 5)
        - Taskmaster: Most missions posted across month (min 5)
        - Legionnaire: Most raid + patrol events across month (min 5)
        - Scout: Most tryouts hosted across month (min 5)
        """
        from database.ac_models import (
            get_hwtm_winner, get_leggionary_winner, 
            get_scout_winner, get_taskmaster_winner, 
            is_last_period_of_month
        )
        from database.models import Member
        
        titles = {}
        
        # ALWAYS calculate HWTM for this period
        hwtm_winner_id, hwtm_count = get_hwtm_winner(self.ac_period)
        if hwtm_winner_id and hwtm_count >= 5:
            winner = Member.query.get(hwtm_winner_id)
            titles['Host with the Most'] = {
                'winner': winner.discord_username if winner else 'Unknown',
                'count': hwtm_count,
                'requirement': '5+ events hosted (Training + Raid + Patrol)',
                'awarded_at_period': self.ac_period.period_name,
                'is_monthly': False
            }
        
        # Only calculate other titles if this is the last period of the month
        if is_last_period_of_month(self.ac_period):
            # Leggionary (most raid + patrol events accumulated, min 5)
            leg_winner_id, leg_count = get_leggionary_winner(self.ac_period)
            if leg_winner_id and leg_count >= 5:
                winner = Member.query.get(leg_winner_id)
                titles['Legionnaire'] = {
                    'winner': winner.discord_username if winner else 'Unknown',
                    'count': leg_count,
                    'requirement': '5+ events hosted (Raids + Patrols across month)',
                    'is_monthly': True
                }
            
            # Scout (most tryouts accumulated, min 5)
            scout_winner_id, scout_count = get_scout_winner(self.ac_period)
            if scout_winner_id and scout_count >= 5:
                winner = Member.query.get(scout_winner_id)
                titles['Scout'] = {
                    'winner': winner.discord_username if winner else 'Unknown',
                    'count': scout_count,
                    'requirement': '5+ tryouts hosted (accumulated across month)',
                    'is_monthly': True
                }
            
            # Taskmaster (most missions accumulated, min 5)
            tm_winner_id, tm_count = get_taskmaster_winner(self.ac_period)
            if tm_winner_id and tm_count >= 5:
                winner = Member.query.get(tm_winner_id)
                titles['Taskmaster'] = {
                    'winner': winner.discord_username if winner else 'Unknown',
                    'count': tm_count,
                    'requirement': '5+ missions posted (accumulated across month)',
                    'is_monthly': True
                }
        
        self.title_winners = titles
        return titles
    
    def generate_title_discord_message(self) -> str:
        """Generate Discord message for title winners"""
        if not self.title_winners:
            return "No title winners this cycle (minimum requirements not met)."
        
        message = f"🏆 **Title Rewards - {self.ac_period.period_name}** 🏆\n\n"
        
        for title, info in self.title_winners.items():
            message += f"**@{title}**\n"
            message += f"👑 Winner: **{info['winner']}**\n"
            message += f"📊 Achievement: {info['count']} ({info['requirement']})\n"
            
            if info.get('is_monthly'):
                message += f"📅 *Monthly Award - Accumulated Stats*\n"
            else:
                message += f"⏱️ *Period Award*\n"
            
            message += "\n"
        
        if len(self.title_winners) < 6:
            monthly_msg = " or end of month has not been reached" if any(info.get('is_monthly') for info in self.title_winners.values()) else ""
            message += f"\n*Note: Some titles had no qualifiers (minimum requirements not met{monthly_msg})*"
        
        return message
    
    def generate_csv_report(self) -> str:
        """Generate CSV report as string"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([f"AC Report - {self.ac_period.period_name}"])
        writer.writerow([f"{self.ac_period.start_date.strftime('%B %d, %Y')} - {self.ac_period.end_date.strftime('%B %d, %Y')}"])
        writer.writerow([])
        
        # Column headers
        writer.writerow(['Discord Username', 'Rank', 'Quota', 'Points Earned', 'Status', 'Total Activities'])
        
        # Data
        for progress in sorted(self.members_progress, key=lambda x: x['member'].discord_username):
            member = progress['member']
            
            if progress['is_protected']:
                status = "PROTECTED (IA)"
            elif progress['points'] >= progress['quota']:
                status = "PASSED"
            else:
                status = "FAILED"
            
            writer.writerow([
                member.discord_username,
                member.current_rank,
                progress['quota'],
                round(progress['points'], 1),
                status,
                len(progress['recent_activities'])
            ])
        
        # Summary
        writer.writerow([])
        writer.writerow(['SUMMARY'])
        total = len(self.members_progress)
        passed = sum(1 for p in self.members_progress if not p['is_protected'] and p['points'] >= p['quota'])
        failed = sum(1 for p in self.members_progress if not p['is_protected'] and p['points'] < p['quota'])
        protected = sum(1 for p in self.members_progress if p['is_protected'])
        
        writer.writerow(['Total Members', total])
        writer.writerow(['Passed', passed])
        writer.writerow(['Failed', failed])
        writer.writerow(['Protected (IA)', protected])
        
        return output.getvalue()


def send_discord_webhook(webhook_url: str, message: str, title: str = "Title Rewards") -> bool:
    """Send title rewards to Discord webhook"""
    import requests
    import json
    
    if not webhook_url:
        return False
    
    embed = {
        "title": title,
        "description": message,
        "color": 0xFFD700,  # Gold color
        "timestamp": datetime.utcnow().isoformat(),
        "footer": {
            "text": "Jedi Taskforce Management System"
        }
    }
    
    payload = {
        "username": "AC Manager",
        "embeds": [embed]
    }
    
    try:
        response = requests.post(
            webhook_url,
            data=json.dumps(payload),
            headers={'Content-Type': 'application/json'},
            timeout=10
        )
        return response.status_code == 204
    except Exception as e:
        print(f"Failed to send webhook: {e}")
        return False