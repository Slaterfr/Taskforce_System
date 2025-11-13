from io import BytesIO
from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database.models import Member
from database.ac_models import ACPeriod, ActivityEntry, InactivityNotice, ACExemption, get_member_quota
from sqlalchemy import func

def _gather_ac_rows(period: ACPeriod) -> List[List]:
    """Return list of rows (including header) for the AC period"""
    header = [
        "Rank",
        "Discord Username",
        "Roblox Username",
        "Quota",
        "Points",
        "Percentage",
        "Status",
        "Protected (IA)",
        "Exempt",
        "Recent Activities"
    ]
    rows = [header]

    if not period:
        return rows

    # Query members with quota
    members = Member.query.filter(Member.is_active == True).order_by(Member.current_rank, Member.discord_username).all()
    for m in members:
        quota = get_member_quota(m.current_rank) or 0
        if quota == 0:
            # skip ranks with no quota
            continue

        total_points = (
            ActivityEntry.query.with_entities(
                ActivityEntry.points
            )
            .filter(ActivityEntry.member_id == m.id, ActivityEntry.ac_period_id == period.id)
            .all()
        )
        total_points = sum(p[0] for p in total_points) if total_points else 0.0

        ia = InactivityNotice.query.filter_by(member_id=m.id, ac_period_id=period.id, protects_ac=True).first()
        exemption = ACExemption.query.filter_by(member_id=m.id, ac_period_id=period.id).first()
        recent_acts = ActivityEntry.query.filter_by(member_id=m.id, ac_period_id=period.id).order_by(ActivityEntry.activity_date.desc()).limit(5).all()
        recent_str = "; ".join(f"{a.activity_date.strftime('%Y-%m-%d')}:{a.activity_type}({a.points})" for a in recent_acts)

        pct = round(min(100.0, (total_points / quota) * 100.0), 2) if quota else 0.0
        
        # Determine status: Exempt takes priority, then IA, then points
        if exemption:
            status = "Exempt"
        elif ia:
            status = "Protected (IA)"
        elif total_points >= quota:
            status = "Passed"
        else:
            status = "In Progress"

        rows.append([
            m.current_rank,
            m.discord_username,
            m.roblox_username or "",
            quota,
            total_points,
            pct,
            status,
            "Yes" if ia else "No",
            "Yes" if exemption else "No",
            recent_str
        ])

    return rows

def _gather_ac_data_by_rank(period: ACPeriod) -> Dict[str, List]:
    """Gather AC data organized by rank"""
    if not period:
        return {}
    
    data_by_rank = {}
    
    # Query members with quota, grouped by rank (exclude Chief General only)
    excluded_ranks = {'chief general'}
    members = Member.query.filter(
        Member.is_active == True,
        func.lower(Member.current_rank).notin_([r.lower() for r in excluded_ranks])
    ).order_by(Member.current_rank, Member.discord_username).all()
    
    for m in members:
        quota = get_member_quota(m.current_rank) or 0
        if quota == 0:
            continue
        
        rank = m.current_rank
        if rank not in data_by_rank:
            data_by_rank[rank] = []
        
        total_points = (
            ActivityEntry.query.with_entities(ActivityEntry.points)
            .filter(ActivityEntry.member_id == m.id, ActivityEntry.ac_period_id == period.id)
            .all()
        )
        total_points = sum(p[0] for p in total_points) if total_points else 0.0
        
        ia = InactivityNotice.query.filter_by(member_id=m.id, ac_period_id=period.id, protects_ac=True).first()
        exemption = ACExemption.query.filter_by(member_id=m.id, ac_period_id=period.id).first()
        
        # Get activity breakdown
        activities = ActivityEntry.query.filter_by(member_id=m.id, ac_period_id=period.id).all()
        activity_counts = {}
        for a in activities:
            activity_counts[a.activity_type] = activity_counts.get(a.activity_type, 0) + 1
        
        pct = round(min(100.0, (total_points / quota) * 100.0), 2) if quota else 0.0
        
        # Determine status
        if exemption:
            status = "Excused"
        elif ia:
            status = "Inactivity"
        elif total_points >= quota:
            status = "Passed"
        else:
            status = "Failed"
        
        data_by_rank[rank].append({
            'member': m,
            'quota': quota,
            'points': total_points,
            'percentage': pct,
            'status': status,
            'ia': bool(ia),
            'exempt': bool(exemption),
            'activity_counts': activity_counts
        })
    
    return data_by_rank

def _write_rows_to_sheet(wb: Workbook, sheet_name: str, rows: List[List], period: ACPeriod = None):
    # create safe unique sheet name
    name = sheet_name[:31]  # Excel limit
    if name in wb.sheetnames:
        # append a suffix
        idx = 1
        base = name
        while f"{base}_{idx}" in wb.sheetnames:
            idx += 1
        name = f"{base}_{idx}"

    ws = wb.create_sheet(title=name)
    
    # Define styles
    dark_purple = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
    light_grey = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Status colors
    status_colors = {
        "Passed": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
        "Failed": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # Light red
        "Inactivity": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light blue
        "Excused": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Yellow
        "Protected (IA)": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "Exempt": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
        "In Progress": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    }
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # If we have period data, use formatted version
    if period:
        data_by_rank = _gather_ac_data_by_rank(period)
        
        # Title row (12 columns: A-L) - use period name
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"Staff Team Activity Checks {period.period_name}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = white_fill
        
        # Period row (12 columns: A-L)
        ws.merge_cells('A2:L2')
        period_cell = ws['A2']
        
        # Format dates with ordinal suffixes
        def ordinal(n):
            return "%d%s" % (n, "th" if 4 <= n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th"))
        
        start_day = period.start_date.day
        end_day = period.end_date.day
        start_str = f"{ordinal(start_day)} of {period.start_date.strftime('%B')}"
        end_str = f"{ordinal(end_day)} of {period.end_date.strftime('%B')}"
        period_cell.value = f"AC CYCLE: {start_str} --> {end_str}"
        period_cell.font = Font(size=12, bold=True)
        period_cell.alignment = Alignment(horizontal='center', vertical='center')
        period_cell.fill = white_fill
        
        # Headers
        headers = ["Rank", "Username", "Result", "Total", "Tryouts", "Events", "Cancelled", 
                   "Evaluations", "Supervisions", "Missions", "IA", "Excused"]
        header_row = 3
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows organized by rank
        row_num = 4
        # Get all ranks from data, ordered by priority (higher ranks first)
        rank_priority = {'Marshal': 1, 'Commander': 2, 'Prospect': 3, 'General': 0, 'Chief General': -1}
        rank_order = sorted(data_by_rank.keys(), 
                          key=lambda r: (rank_priority.get(r, 999), r))
        
        for rank in rank_order:
            members_data = data_by_rank[rank]
            
            # Rank header row (12 columns: A-L)
            ws.merge_cells(f'A{row_num}:L{row_num}')
            rank_cell = ws.cell(row=row_num, column=1, value=rank)
            rank_cell.fill = dark_purple
            rank_cell.font = white_font
            rank_cell.alignment = Alignment(horizontal='left', vertical='center')
            rank_cell.border = border
            row_num += 1
            
            # Member rows
            for member_data in members_data:
                m = member_data['member']
                activity_counts = member_data['activity_counts']
                
                # Determine row color (alternating)
                row_fill = light_grey if (row_num - header_row) % 2 == 0 else white_fill
                
                # Rank
                cell = ws.cell(row=row_num, column=1, value=rank)
                cell.fill = row_fill
                cell.border = border
                
                # Username
                cell = ws.cell(row=row_num, column=2, value=m.discord_username)
                cell.fill = row_fill
                cell.border = border
                
                # Result (Status) - color coded
                status = member_data['status']
                cell = ws.cell(row=row_num, column=3, value=status)
                cell.fill = status_colors.get(status, white_fill)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Total (Points)
                cell = ws.cell(row=row_num, column=4, value=round(member_data['points'], 1))
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Activity counts - map to columns: Tryouts, Events, Cancelled, Evaluations, Supervisions, Missions
                # Tryouts
                tryout_count = activity_counts.get('Tryout', 0)
                cell = ws.cell(row=row_num, column=5, value=tryout_count if tryout_count > 0 else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Events (Raid + Patrol combined)
                events_count = activity_counts.get('Raid', 0) + activity_counts.get('Patrol', 0)
                cell = ws.cell(row=row_num, column=6, value=events_count if events_count > 0 else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Cancelled (Canceled Training + Cancelled Tryout)
                cancelled_count = activity_counts.get('Canceled Training', 0) + activity_counts.get('Cancelled Tryout', 0)
                cell = ws.cell(row=row_num, column=7, value=cancelled_count if cancelled_count > 0 else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Evaluations
                eval_count = activity_counts.get('Evaluation', 0)
                cell = ws.cell(row=row_num, column=8, value=eval_count if eval_count > 0 else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Supervisions
                super_count = activity_counts.get('Supervision', 0)
                cell = ws.cell(row=row_num, column=9, value=super_count if super_count > 0 else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Missions
                mission_count = activity_counts.get('Mission', 0)
                cell = ws.cell(row=row_num, column=10, value=mission_count if mission_count > 0 else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # IA checkbox (Yes/No)
                cell = ws.cell(row=row_num, column=11, value="✓" if member_data['ia'] else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Excused checkbox (Yes/No)
                cell = ws.cell(row=row_num, column=12, value="✓" if member_data['exempt'] else "")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                row_num += 1
            
            # Total row for rank (optional - can add if needed)
            # row_num += 1
        
        # Set column widths (12 columns)
        column_widths = [12, 20, 12, 8, 8, 8, 10, 12, 12, 10, 5, 8]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    else:
        # Fallback to simple version if no period
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Auto-adjust column widths
        for i, _ in enumerate(rows[0], start=1):
            col = get_column_letter(i)
            max_len = max(
                (len(str(cell)) if cell is not None else 0)
                for cell in (r[i-1] for r in rows)
            )
            ws.column_dimensions[col].width = min(max(10, max_len + 2), 60)

def generate_ac_workbook_bytes(period_id: int = None) -> (BytesIO, str):
    """
    Build a new workbook for given AC period_id (or active if None).
    Returns (BytesIO, filename).
    """
    if period_id:
        period = ACPeriod.query.get(period_id)
    else:
        period = ACPeriod.query.filter_by(is_active=True).first()

    period_name = period.period_name if period else f"AC_{datetime.utcnow().strftime('%Y%m%d')}"
    rows = _gather_ac_rows(period)

    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)
    _write_rows_to_sheet(wb, f"AC_{period_name}", rows, period=period)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"AC_{period_name}.xlsx"
    return out, filename

def merge_into_uploaded_workbook_bytes(uploaded_file_stream, period_id: int = None) -> (BytesIO, str):
    """
    Load uploaded workbook (file-like), append a sheet with the AC data, and return bytes.
    """
    # load existing
    uploaded_file_stream.seek(0)
    try:
        wb = load_workbook(uploaded_file_stream)
    except Exception:
        # if load fails, create new workbook
        wb = Workbook()
        wb.remove(wb.active)

    if period_id:
        period = ACPeriod.query.get(period_id)
    else:
        period = ACPeriod.query.filter_by(is_active=True).first()

    period_name = period.period_name if period else f"AC_{datetime.utcnow().strftime('%Y%m%d')}"
    rows = _gather_ac_rows(period)
    _write_rows_to_sheet(wb, f"AC_{period_name}", rows, period=period)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"merged_AC_{period_name}.xlsx"
    return out, filename